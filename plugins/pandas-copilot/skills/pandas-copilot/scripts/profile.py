#!/usr/bin/env python3
"""pandas-copilot Stage 2 profiler — deterministic data profiling.

Usage:
    <venv-python> profile.py <data-file> [--sheet NAME] [--nrows N] [--head N]

Reads the data file (xlsx/xls/csv/json/parquet), runs an Excel health check
(sheets, merged cells, hidden rows/columns, header suspicion, type traps) and
prints a per-table profile: shape, per-column dtype/nulls/uniques/samples,
head rows, duplicate count. Read-only: never modifies the input file.

Output is a plain-text report on stdout, meant to be interpreted and relayed
to the user. Exit codes: 0 = report produced, 1 = could not read the file.
"""

import argparse
import os
import re
import sys

import pandas as pd

MAX_COLS_SHOWN = 40
SAMPLE_VALUES = 3
EXCEL_EXTS = {".xlsx", ".xlsm", ".xls"}


def fmt_count(n):
    return f"{n:,}"


def truncate(text, width=24):
    text = str(text)
    return text if len(text) <= width else text[: width - 1] + "…"


def scalar(v):
    """numpy scalar -> python scalar so repr reads clean (1, not np.int64(1))."""
    return v.item() if hasattr(v, "item") else v


# ---------------------------------------------------------------- health check

def excel_health_check(path, sheet_frames):
    """Structural + type-trap diagnosis. Returns a list of issue strings."""
    issues = []

    if len(sheet_frames) > 1:
        issues.append(
            f"Multi-sheet file ({len(sheet_frames)} sheets): "
            + ", ".join(repr(s) for s in sheet_frames)
            + " — confirm which sheet(s) to process."
        )

    # Structural checks need openpyxl and only work on .xlsx/.xlsm.
    if os.path.splitext(path)[1].lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=False)
            for name in wb.sheetnames:
                ws = wb[name]
                merged = list(ws.merged_cells.ranges)
                if merged:
                    shown = ", ".join(str(r) for r in merged[:5])
                    more = f" (+{len(merged) - 5} more)" if len(merged) > 5 else ""
                    issues.append(
                        f"Sheet {name!r}: {len(merged)} merged region(s): {shown}{more}"
                        " — values live only in the top-left cell; expect NaN + ffill."
                    )
                hidden_rows = [r for r in ws.row_dimensions if ws.row_dimensions[r].hidden]
                hidden_cols = [c for c in ws.column_dimensions if ws.column_dimensions[c].hidden]
                if hidden_rows or hidden_cols:
                    issues.append(
                        f"Sheet {name!r}: hidden rows {hidden_rows[:10]} / hidden columns "
                        f"{hidden_cols[:10]} — pandas reads them anyway; confirm whether to keep."
                    )
        except Exception as exc:  # structural check is best-effort
            issues.append(f"openpyxl structural check skipped: {exc}")
    elif os.path.splitext(path)[1].lower() == ".xls":
        issues.append("Legacy .xls file: merged-cell/hidden-row checks unavailable (openpyxl cannot read .xls).")

    for name, df in sheet_frames.items():
        prefix = f"Sheet {name!r}: " if len(sheet_frames) > 1 else ""

        unnamed = [c for c in df.columns if str(c).startswith("Unnamed:")]
        if unnamed:
            issues.append(
                f"{prefix}{len(unnamed)} 'Unnamed:' column header(s) — likely title rows or a "
                "multi-row header; consider read params `header=`/`skiprows=`."
            )

        for col in df.columns:
            s = df[col].dropna()
            if s.empty:
                continue
            if pd.api.types.is_object_dtype(s) or pd.api.types.is_string_dtype(s):
                sv = s.astype(str)
                if sv.str.fullmatch(r"[45]\d{4}").mean() > 0.8:
                    issues.append(
                        f"{prefix}column {col!r} looks like Excel date serial numbers stored as "
                        "text (e.g. 44197 = 2021-01-01)."
                    )
                if sv.str.contains(r"\d,\d{3}").any():
                    issues.append(
                        f"{prefix}column {col!r} contains thousands separators — numeric values "
                        "stored as text; clean before arithmetic."
                    )
                if sv.str.fullmatch(r"0\d+").any():
                    issues.append(
                        f"{prefix}column {col!r} has values with leading zeros — must be read "
                        f"as str (`dtype={{{col!r}: str}}`) or the zeros are lost."
                    )
            elif pd.api.types.is_integer_dtype(s) or pd.api.types.is_float_dtype(s):
                if pd.api.types.is_float_dtype(s) and (s % 1 == 0).all() and s.abs().max() >= 1e6:
                    issues.append(
                        f"{prefix}column {col!r} is float but all-integral with large values — "
                        "possible ID/phone read as number (precision + leading-zero risk); "
                        "consider reading as str."
                    )
                if pd.api.types.is_integer_dtype(s) and s.between(20000, 60000).mean() > 0.8 and s.nunique() > 5:
                    issues.append(
                        f"{prefix}column {col!r} is integers in the 20000–60000 range — may be "
                        "Excel date serials (convert with origin='1899-12-30', unit='D')."
                    )
    return issues


# --------------------------------------------------------------------- readers

def read_csv_robust(path, nrows):
    last_err = None
    for enc in ("utf-8-sig", "gb18030", "latin-1"):
        try:
            return {"<csv>": pd.read_csv(path, encoding=enc, nrows=nrows)}, enc
        except UnicodeDecodeError as exc:
            last_err = exc
        except pd.errors.ParserError:
            # delimiter trouble — let the python engine sniff it
            return {"<csv>": pd.read_csv(path, encoding=enc, nrows=nrows, sep=None, engine="python")}, f"{enc} (sniffed sep)"
    raise last_err


def read_any(path, sheet, nrows):
    """Returns ({table_name: DataFrame}, note) — Excel maps every sheet."""
    ext = os.path.splitext(path)[1].lower()
    if ext in EXCEL_EXTS:
        xlsx = pd.ExcelFile(path)
        names = [sheet] if sheet is not None else xlsx.sheet_names
        return {n: xlsx.parse(n, nrows=nrows) for n in names}, None
    if ext == ".csv":
        return read_csv_robust(path, nrows)
    if ext == ".json":
        try:
            return {"<json>": pd.read_json(path)}, None
        except ValueError:
            return {"<json>": pd.read_json(path, lines=True)}, "read as JSON Lines"
    if ext == ".parquet":
        return {"<parquet>": pd.read_parquet(path)}, None
    raise ValueError(f"Unsupported extension {ext!r} (supported: xlsx/xls/csv/json/parquet)")


# --------------------------------------------------------------------- profile

def profile_table(name, df, head_n):
    print(f"\n## Table: {name}")
    print(f"shape: {fmt_count(len(df))} rows x {len(df.columns)} columns")
    try:
        dup = int(df.duplicated().sum())
        print(f"fully duplicated rows: {fmt_count(dup)}")
    except TypeError:
        print("fully duplicated rows: n/a (unhashable cell values)")
    print(f"memory: {df.memory_usage(deep=True).sum() / 1024:.1f} KB")

    print("\n### Columns (name | dtype | non-null | null% | unique | sample values)")
    for col in df.columns[:MAX_COLS_SHOWN]:
        s = df[col]
        null_pct = 100.0 * s.isna().mean() if len(s) else 0.0
        try:
            uniq = fmt_count(s.nunique(dropna=True))
        except TypeError:
            uniq = "n/a"
        samples = ", ".join(truncate(repr(scalar(v))) for v in s.dropna().unique()[:SAMPLE_VALUES])
        print(f"- {col!r} | {s.dtype} | {fmt_count(s.notna().sum())} | {null_pct:.1f}% | {uniq} | {samples}")
    if len(df.columns) > MAX_COLS_SHOWN:
        print(f"... ({len(df.columns) - MAX_COLS_SHOWN} more columns not shown)")

    print(f"\n### Head (first {min(head_n, len(df))} rows)")
    with pd.option_context("display.max_columns", MAX_COLS_SHOWN, "display.width", 200):
        print(df.head(head_n).to_string())


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("data_file")
    ap.add_argument("--sheet", default=None, help="profile only this Excel sheet")
    ap.add_argument("--nrows", type=int, default=None, help="read at most N rows (large files)")
    ap.add_argument("--head", type=int, default=5, help="head rows to show (default 5)")
    args = ap.parse_args()

    path = args.data_file
    if not os.path.exists(path):
        print(f"ERROR: file not found: {path}", file=sys.stderr)
        return 1

    print(f"# Data Profile: {path}")
    print(f"file size: {os.path.getsize(path) / 1024:.1f} KB")

    try:
        frames, note = read_any(path, args.sheet, args.nrows)
    except ImportError as exc:
        print(f"ERROR: missing engine for this format: {exc}\n"
              "Install it into the venv (e.g. `uv pip install --python <venv-python> pyarrow`).",
              file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"ERROR: could not read file: {type(exc).__name__}: {exc}\n"
              "For Excel traps (multi-row headers, merged cells, engines) see excel-gotchas.md.",
              file=sys.stderr)
        return 1

    if note:
        print(f"read note: {note}")
    if args.nrows:
        print(f"NOTE: profiling only the first {args.nrows} rows — totals/duplicates are partial.")

    if os.path.splitext(path)[1].lower() in EXCEL_EXTS:
        print("\n## Excel health check")
        issues = excel_health_check(path, frames)
        if issues:
            for issue in issues:
                print(f"- [!] {issue}")
            print("(mitigations for each trap: see excel-gotchas.md)")
        else:
            print("- no common Excel issues detected")

    for name, df in frames.items():
        profile_table(name, df, args.head)

    print("\n(profile complete — interpret findings and present them to the user before Gate A)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
