# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "openpyxl>=3.1",
# ]
# ///
"""Build the final three-sheet PD rule workbook from .pd-extraction/ artifacts.

Sheets:
  1. PD规则表  - 14 fixed columns, sorted by rule_id, frozen header
  2. 类别统计  - rule counts per category (desc) + total row
  3. 说明      - key/value metadata: tool id, timestamp, inputs, column
                 definitions, severity source note, exhaustiveness evidence

Exit codes: 0 success, 1 input error (missing/empty/corrupt data).
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_CATEGORIES = [
    "入选排除", "知情同意", "访视", "给药", "样本采集", "检查评估",
    "合并用药", "不良事件", "随机盲法", "试验药品", "程序",
]

# (header, jsonl field, column width)
RULE_COLUMNS = [
    ("规则编号", "rule_id", 14),
    ("PD类别", "category", 12),
    ("PD子类", "subcategory", 16),
    ("规则描述", "description", 48),
    ("判定逻辑", "condition", 48),
    ("相关访视", "visits", 16),
    ("相关表单", "forms", 16),
    ("相关字段/数据点", "fields", 22),
    ("来源文件", "source_files", 26),
    ("来源定位", "source_locator", 30),
    ("原始规则编号", "original_rule_ids", 18),
    ("严重程度", "severity", 12),
    ("重复规则指向", "duplicate_of", 16),
    ("备注", "remarks", 36),
]

REQUIRED_RULE_FIELDS = {"rule_id", "category", "description"}

COLUMN_DEFINITIONS = [
    ("规则编号", "本表统一编号（PD-xxx），提取阶段分配，全表唯一"),
    ("PD类别", "11 个固定类别之一（入选排除/知情同意/访视/给药/样本采集/检查评估/合并用药/不良事件/随机盲法/试验药品/程序）"),
    ("PD子类", "类别下的细分主题，由提取阶段按源文档内容归纳"),
    ("规则描述", "方案偏离的判定规则，以可执行核查的语言表述"),
    ("判定逻辑", "SQL 风格形式化表达式（数据集.变量 全限定、比较符、EXISTS/NOT EXISTS、DATEDIFF 等），可直接作为编写偏离数据拉取代码的依据"),
    ("相关访视", "规则适用的访视（多个以逗号分隔；全程适用写'所有访视'；空表示不限访视）"),
    ("相关表单", "规则涉及的 CRF 表单名（以数据库设计文件为准；未知或不适用留空）"),
    ("相关字段/数据点", "判定逻辑引用的变量（数据集.变量 全限定，逗号分隔，与判定逻辑中出现的变量一致）"),
    ("来源文件", "规则来源的文档（多个来源以分号分隔，表示多源合并）"),
    ("来源定位", "章节号 / Sheet名+行范围 / 页码，用于溯源到原文"),
    ("原始规则编号", "源文档中的原始规则编号（如 DVP 规则号；多源合并时全部保留）"),
    ("严重程度", "偏离分级（如 Major/Minor）；源文档未定义分级标准时留空"),
    ("重复规则指向", "跨来源合并时，本行（主规则）记录被合并重复项的原始规则编号或来源定位；未发生合并留空"),
    ("备注", "补充说明；审核未通过的规则在此标注原因"),
]

TOOL_ID = "pd-rule-extractor skill"


def eprint(msg: str) -> None:
    print(msg, file=sys.stderr)


def fail(msg: str, code: int = 1) -> None:
    eprint(f"error: {msg}")
    sys.exit(code)


def load_json(path: Path, what: str):
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        fail(f"corrupt {what}: {path} ({exc})")


def load_rules(path: Path) -> list[dict]:
    if not path.exists():
        fail(f"missing rules-written.jsonl: {path}")
    rules: list[dict] = []
    for lineno, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            rules.append(json.loads(line))
        except json.JSONDecodeError as exc:
            fail(f"corrupt rules-written.jsonl line {lineno}: {exc}")
    if not rules:
        fail(f"rules-written.jsonl is empty: {path} (no rules to write)")
    for i, r in enumerate(rules, start=1):
        missing = REQUIRED_RULE_FIELDS - set(r)
        if missing:
            fail(f"rules-written.jsonl line {i}: missing required fields {sorted(missing)}")
    return rules


def as_text(value) -> str:
    """Render a JSONL field as cell text: lists join with '; ', None -> ''."""
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value if v is not None)
    if isinstance(value, (dict, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def style_header(ws, ncols: int) -> None:
    font = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9E1F2")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def build_rules_sheet(ws, rules: list[dict]) -> None:
    for ci, (header, _field, width) in enumerate(RULE_COLUMNS, start=1):
        ws.cell(row=1, column=ci, value=header)
        ws.column_dimensions[get_column_letter(ci)].width = width
    style_header(ws, len(RULE_COLUMNS))

    wrap = Alignment(vertical="top", wrap_text=True)
    for ri, rule in enumerate(sorted(rules, key=lambda r: str(r.get("rule_id", ""))), start=2):
        remarks = as_text(rule.get("remarks"))
        if rule.get("review_status") == "failed":
            notes = as_text(rule.get("review_notes"))
            note_text = f"审核未通过：{notes}" if notes else "审核未通过"
            remarks = f"{remarks}；{note_text}" if remarks else note_text
        row_values = []
        for _header, field, _w in RULE_COLUMNS:
            if field == "remarks":
                row_values.append(remarks)
            else:
                row_values.append(as_text(rule.get(field)))
        for ci, value in enumerate(row_values, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.alignment = wrap


def build_stats_sheet(ws, rules: list[dict], categories: list[str]) -> None:
    ws.cell(row=1, column=1, value="PD类别")
    ws.cell(row=1, column=2, value="规则数")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    style_header(ws, 2)

    counts: dict[str, int] = {}
    for r in rules:
        cat = str(r.get("category", "") or "(未分类)")
        counts[cat] = counts.get(cat, 0) + 1
    # Fixed category order first (any present), then non-standard ones.
    ordered = [c for c in categories if c in counts]
    ordered += sorted(c for c in counts if c not in categories)
    ordered.sort(key=lambda c: -counts[c])

    ri = 2
    for cat in ordered:
        ws.cell(row=ri, column=1, value=cat)
        ws.cell(row=ri, column=2, value=counts[cat])
        ri += 1
    total_cell = ws.cell(row=ri, column=1, value="合计")
    total_cell.font = Font(bold=True)
    sum_cell = ws.cell(row=ri, column=2, value=sum(counts.values()))
    sum_cell.font = Font(bold=True)


def build_notes_sheet(ws, *, generated_at: str, inputs_meta: str, rules: list[dict],
                      coverage: dict | None, reconciliation: dict | None,
                      severity_status: str, loop_rounds: int | None) -> None:
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100
    wrap = Alignment(vertical="top", wrap_text=True)
    key_font = Font(bold=True)

    rows: list[tuple[str, str]] = []
    n_passed = sum(1 for r in rules if r.get("review_status") == "passed")
    n_failed = sum(1 for r in rules if r.get("review_status") == "failed")

    rows.append(("生成工具", TOOL_ID))
    rows.append(("生成时间", generated_at))
    rows.append(("规则总数", str(len(rules))))
    rows.append(("输入文件", inputs_meta))

    rows.append(("严重程度来源", SEVERITY_NOTES.get(severity_status, SEVERITY_NOTES["undefined"])))

    # Exhaustiveness evidence: coverage + reconciliation + review summary.
    if coverage:
        cs = coverage.get("summary", {})
        doc_bits = []
        matrix = coverage.get("matrix", {})
        for doc_id in coverage.get("documents", []):
            cats = matrix.get(doc_id, {})
            n_cov = sum(1 for c in cats.values() if c["status"] == "covered")
            n_none = sum(1 for c in cats.values() if c["status"] == "none-confirmed")
            n_mis = sum(1 for c in cats.values() if c["status"] == "missing")
            doc_bits.append(f"{doc_id}: 有规则类别 {n_cov}，确认无类别 {n_none}，缺失类别 {n_mis}")
        rows.append(("穷尽性证据-覆盖矩阵",
                     f"覆盖 {cs.get('covered', 0)} 格 / 确认无 {cs.get('none_confirmed', 0)} 格 / "
                     f"缺失 {cs.get('missing', 0)} 格。" + "；".join(doc_bits)))
    else:
        rows.append(("穷尽性证据-覆盖矩阵", "未提供 coverage.json"))

    if reconciliation:
        dvp = reconciliation.get("dvp", {})
        rows.append(("穷尽性证据-DVP对账",
                     f"DVP 规则共 {dvp.get('total_rules', 0)} 条：映射 {dvp.get('mapped', 0)} 条 / "
                     f"显式排除(non-PD) {dvp.get('non_pd', 0)} 条 / 缺口 "
                     f"{dvp.get('unmapped_gaps', 0) + dvp.get('dangling_ref_gaps', 0)} 条"))
        units = reconciliation.get("units", {})
        rows.append(("穷尽性证据-章节引用",
                     f"文档单元 {units.get('referenced', 0)}/{units.get('total', 0)} 被规则引用，"
                     f"{units.get('unreferenced', 0)} 个单元未被引用（候选缺口见 reconciliation.json）"))
    else:
        rows.append(("穷尽性证据-DVP对账", "未提供 reconciliation.json"))

    if loop_rounds is not None:
        rounds_text = f"{loop_rounds} 轮（连续两轮无新增收敛）"
    else:
        rounds_text = "未记录（未提供 --loop-rounds）"
    rows.append(("穷尽性证据-补提循环", rounds_text))

    rows.append(("审核汇总", f"通过 {n_passed} 条；未通过 {n_failed} 条（已在备注列标注）"))

    for header, definition in COLUMN_DEFINITIONS:
        rows.append((f"字段释义-{header}", definition))

    for ri, (key, value) in enumerate(rows, start=1):
        kcell = ws.cell(row=ri, column=1, value=key)
        kcell.font = key_font
        kcell.alignment = wrap
        vcell = ws.cell(row=ri, column=2, value=value)
        vcell.alignment = wrap


def severity_criteria_status(path: Path) -> str:
    """Classify severity-criteria.md as 'override' | 'defined' | 'undefined'.

    - 'undefined': file missing/unreadable, or its first non-empty line records
      "未定义" (stage 0 writes that marker when source documents define nothing)
    - 'override': file mentions severity_override (stage 3 writes the override
      declaration into this file when the project config forces one value)
    - 'defined': otherwise (criteria extracted from project documents)
    """
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return "undefined"
    first = next((ln for ln in text.splitlines() if ln.strip()), "")
    if "未定义" in first:
        return "undefined"
    if "severity_override" in text:
        return "override"
    return "defined"


SEVERITY_NOTES = {
    "override": "由项目配置 severity_override 统一覆盖（详见 severity-criteria.md）",
    "defined": "按项目文档定义（severity-criteria.md）",
    "undefined": "源文档未定义严重程度标准，该列留空",
}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--workdir", required=True, help="Working directory (.pd-extraction).")
    parser.add_argument("--output", required=True, help="Output xlsx path.")
    parser.add_argument("--inputs-meta", required=True,
                        help="Semicolon-separated input file names shown on the notes sheet.")
    parser.add_argument("--generated-at", default=None,
                        help="Timestamp string (YYYY-MM-DD HH:MM:SS). Defaults to now.")
    parser.add_argument("--categories", default=None,
                        help="Comma-separated category list overriding the default 11 (debug use only; the category literals are fixed by contract).")
    parser.add_argument("--loop-rounds", type=int, default=None,
                        help="Number of gap-fix loop rounds executed in stage 5 "
                             "(recorded on the notes sheet as exhaustiveness evidence).")
    args = parser.parse_args()

    workdir = Path(args.workdir)
    if not workdir.is_dir():
        fail(f"workdir not found: {workdir}")

    generated_at = args.generated_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if args.generated_at:
        try:
            datetime.strptime(args.generated_at, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            fail(f"--generated-at must match YYYY-MM-DD HH:MM:SS, got {args.generated_at!r}")

    categories = ([c.strip() for c in args.categories.split(",") if c.strip()]
                  if args.categories else DEFAULT_CATEGORIES)

    rules = load_rules(workdir / "rules-written.jsonl")
    coverage = load_json(workdir / "coverage.json", "coverage.json")
    reconciliation = load_json(workdir / "reconciliation.json", "reconciliation.json")

    severity_status = severity_criteria_status(workdir / "severity-criteria.md")

    wb = Workbook()
    ws_rules = wb.active
    ws_rules.title = "PD规则表"
    build_rules_sheet(ws_rules, rules)

    ws_stats = wb.create_sheet("类别统计")
    build_stats_sheet(ws_stats, rules, categories)

    ws_notes = wb.create_sheet("说明")
    build_notes_sheet(ws_notes, generated_at=generated_at, inputs_meta=args.inputs_meta,
                      rules=rules, coverage=coverage, reconciliation=reconciliation,
                      severity_status=severity_status, loop_rounds=args.loop_rounds)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(out_path))
    except Exception as exc:
        fail(f"failed to write {out_path}: {exc}")

    json.dump({"output": str(out_path), "rules": len(rules),
               "sheets": wb.sheetnames, "generated_at": generated_at},
              sys.stdout, ensure_ascii=False, indent=2)
    print()
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except SystemExit:
        raise
    except Exception as exc:
        fail(f"{type(exc).__name__}: {exc}")
