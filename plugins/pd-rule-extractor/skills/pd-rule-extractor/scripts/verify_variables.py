# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "openpyxl>=3.1",
# ]
# ///
"""Verify that every dataset.variable referenced by rule conditions and fields
actually exists in the source documents.

Index source: only role="dbdesign" documents contribute. Two xlsx layouts are
supported — per-form sheets (sheet name = dataset/form, header row = variables,
headers extracted deterministically by probe_docs.py into doc-map.json) and
dictionary sheets (Dataset/Variable columns, one row per variable; rows are
re-read from the source xlsx, so the path recorded in doc-map.json must remain
accessible). References are extracted from rules-written.jsonl `condition` and
`fields` with regexes.

Outputs variable-verification.json:
  index        sheets/columns counts used for verification
  references   unique dataset.variable refs found (plus EXISTS dataset refs)
  resolved     refs confirmed against the index
  unresolved   refs NOT found in the index (must be fixed before review)
  unverifiable docs have no xlsx index at all (non-xlsx DB design); refs to
               datasets that cannot be checked deterministically

Exit codes: 0 all referenced variables verified (or nothing to check),
            1 hard input error, 2 unresolved references remain.

Usage:
  uv run verify_variables.py --workdir .pd-extraction [--summary]
"""

from __future__ import annotations

import argparse
import difflib
import json
import re
import sys
from pathlib import Path
from typing import NoReturn

# dataset.variable references, e.g. DM.AGE, VISIT.VISDAT, EditChecks.规则编号.
# Both parts must start with a non-digit word char (Unicode-aware, so Chinese
# sheet/column names match) — this also excludes decimals and version literals
# like 'v3.0' (their post-dot part starts with a digit).
REF_RE = re.compile(r"\b([^\W\d]\w*)\.([^\W\d]\w*)\b")
# bare dataset inside EXISTS/NOT EXISTS ( ... )
EXISTS_RE = re.compile(r"(?:NOT\s+)?EXISTS\s*\(\s*([^\W\d]\w*)", re.IGNORECASE)


def eprint(msg: str) -> None:
    print(msg, file=sys.stderr)


def fail(msg: str, code: int = 1) -> NoReturn:
    eprint(f"error: {msg}")
    sys.exit(code)


def norm(name: str) -> str:
    """Normalization for dataset/sheet matching: case- and separator-insensitive."""
    return re.sub(r"[\s_\-]+", "", name).lower()


# Header names marking a dictionary-layout sheet (one row per variable):
# a column holding the dataset name + a column holding the variable name.
DATASET_HEADER_CANDIDATES = {"dataset", "数据集", "ds", "table", "表", "数据表"}
VAR_HEADER_CANDIDATES = {"variable", "变量", "var", "field", "字段", "列", "变量名"}


def load_json(path: Path, what: str):
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        fail(f"corrupt {what}: {path} ({exc})")


def load_jsonl(path: Path, what: str) -> list[dict]:
    if not path.exists():
        fail(f"missing {what}: {path}")
    rows: list[dict] = []
    for lineno, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError as exc:
            fail(f"corrupt {what} line {lineno}: {exc}")
    return rows


def _detect_dictionary_columns(headers: list[str]) -> tuple[int, int] | None:
    """Dictionary layout has a dataset-name column and a variable-name column.
    Returns (ds_idx, var_idx) or None for non-dictionary sheets."""
    norm_headers = [norm(h) for h in headers]
    ds_candidates = {norm(c) for c in DATASET_HEADER_CANDIDATES}
    var_candidates = {norm(c) for c in VAR_HEADER_CANDIDATES}
    ds_idx = next((i for i, h in enumerate(norm_headers) if h in ds_candidates), None)
    var_idx = next((i for i, h in enumerate(norm_headers) if h in var_candidates), None)
    if ds_idx is None or var_idx is None or ds_idx == var_idx:
        return None
    return ds_idx, var_idx


def _index_dictionary_sheet(xlsx_path: Path, sheet_name: str,
                            cols: tuple[int, int], doc_id, index: dict) -> None:
    """Row-based dictionary layout (Dataset | Variable | Label ...): read the
    source xlsx and index dataset names from row values."""
    ds_idx, var_idx = cols
    from openpyxl import load_workbook
    try:
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as exc:
        eprint(f"warning: cannot open {xlsx_path} for dictionary sheet {sheet_name}: {exc}")
        return
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header_row_idx = next(
        (i for i, r in enumerate(rows) if any(v is not None and str(v).strip() for v in r)),
        None,
    )
    if header_row_idx is None:
        wb.close()
        return
    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        ds = row[ds_idx] if ds_idx < len(row) else None
        var = row[var_idx] if var_idx < len(row) else None
        if ds is None or var is None:
            continue
        ds, var = str(ds).strip(), str(var).strip()
        if not ds or not var:
            continue
        entry = index.setdefault(norm(ds), {
            "sheet": sheet_name, "doc_id": doc_id, "columns": {}, "layout": "dictionary",
        })
        entry["columns"].setdefault(norm(var), var)
    wb.close()


def build_index(doc_map: dict) -> dict[str, dict]:
    """sheet/dataset-name (normalized) -> {sheet, doc_id, columns: {norm: original}}.

    Only DB-design documents (role == "dbdesign") contribute — DVP xlsx sheets
    must not pollute the index (they would turn a missing DB-design index into
    spurious 'dataset not in index' failures). Two xlsx layouts are supported:
    per-form sheets (sheet name = dataset/form, header row = variables) and
    dictionary sheets (one row per variable with Dataset/Variable columns).
    """
    index: dict[str, dict] = {}
    for doc in doc_map.get("documents", []):
        if doc.get("role") != "dbdesign":
            continue
        doc_path = Path(doc.get("path", ""))
        for sheet in doc.get("sheets", []):
            headers = [str(h).strip() for h in sheet.get("headers", []) if str(h).strip()]
            if not headers:
                continue
            dict_cols = _detect_dictionary_columns(headers) if doc.get("format") == "xlsx" else None
            if dict_cols is not None:
                if doc_path.exists():
                    _index_dictionary_sheet(doc_path, sheet.get("sheet", ""), dict_cols,
                                            doc.get("doc_id"), index)
                else:
                    # Dictionary rows live in the source file, not doc-map; without
                    # it this sheet contributes nothing (fail-closed: refs to its
                    # datasets stay unresolved). Never fall back to form layout here —
                    # the dictionary headers are NOT variable names.
                    eprint(f"warning: dictionary sheet '{sheet.get('sheet', '')}' needs the "
                           f"source xlsx for row values, but it is not accessible: {doc_path}; "
                           f"its variables will be reported unresolved")
                continue
            # Form layout: sheet name is the dataset/form, headers are variables.
            cols = {}
            for h in headers:
                cols.setdefault(norm(h), h)
            entry = index.setdefault(norm(sheet.get("sheet", "")), {
                "sheet": sheet.get("sheet", ""),
                "doc_id": doc.get("doc_id"),
                "columns": {},
                "layout": "form",
            })
            entry["doc_id"] = doc.get("doc_id")
            entry["columns"].update(cols)
    return index


def extract_refs(rule: dict) -> list[dict]:
    """All dataset.variable refs (condition + fields) and EXISTS dataset refs."""
    refs: list[dict] = []
    seen: set[tuple[str, str]] = set()
    texts = [str(rule.get("condition", "") or ""), str(rule.get("fields", "") or "")]
    for text in texts:
        for m in REF_RE.finditer(text):
            key = (m.group(1), m.group(2))
            if key not in seen:
                seen.add(key)
                refs.append({"kind": "variable", "dataset": key[0], "variable": key[1],
                             "ref": f"{key[0]}.{key[1]}"})
        for m in EXISTS_RE.finditer(text):
            ds = m.group(1)
            key = (ds, "")
            if key not in seen:
                seen.add(key)
                refs.append({"kind": "dataset", "dataset": ds, "variable": None, "ref": ds})
    return refs


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--workdir", required=True, help="Working directory (.pd-extraction).")
    parser.add_argument("--summary", action="store_true",
                        help="Print human-readable summary (still writes JSON output).")
    args = parser.parse_args()

    workdir = Path(args.workdir)
    if not workdir.is_dir():
        fail(f"workdir not found: {workdir}")

    doc_map = load_json(workdir / "doc-map.json", "doc-map.json")
    if doc_map is None:
        fail(f"missing doc-map.json: {workdir / 'doc-map.json'} (run probe_docs.py extract first)")
    rules = load_jsonl(workdir / "rules-written.jsonl", "rules-written.jsonl")

    index = build_index(doc_map)
    has_index = bool(index)

    resolved: list[dict] = []
    unresolved: list[dict] = []
    n_refs = 0
    for rule in rules:
        rid = rule.get("rule_id", "?")
        for ref in extract_refs(rule):
            n_refs += 1
            if not has_index:
                continue  # counted as unverifiable below
            entry = index.get(norm(ref["dataset"]))
            if entry is None:
                suggestion = difflib.get_close_matches(norm(ref["dataset"]), index.keys(), n=3)
                unresolved.append({**ref, "rule_id": rid, "problem": "dataset not in index",
                                   "candidates": [index[s]["sheet"] for s in suggestion]})
                continue
            if ref["kind"] == "dataset":
                resolved.append({**ref, "rule_id": rid, "sheet": entry["sheet"]})
                continue
            if norm(ref["variable"]) not in entry["columns"]:
                suggestion = difflib.get_close_matches(
                    norm(ref["variable"]), entry["columns"].keys(), n=3)
                unresolved.append({**ref, "rule_id": rid,
                                   "problem": f"variable not in sheet {entry['sheet']}",
                                   "candidates": [entry["columns"][s] for s in suggestion]})
                continue
            resolved.append({**ref, "rule_id": rid, "sheet": entry["sheet"]})

    result = {
        "index": {"sheets": len(index),
                  "columns_total": sum(len(e["columns"]) for e in index.values()),
                  "available": has_index},
        "references_total": n_refs,
        "resolved_count": len(resolved),
        "unresolved_count": len(unresolved) if has_index else 0,
        "unverifiable_count": 0 if has_index else n_refs,
        "resolved": resolved,
        "unresolved": unresolved if has_index else [],
        "note": ("no xlsx sheet index in doc-map.json; references cannot be verified "
                 "deterministically (DB design doc is not xlsx) — verify via LLM against "
                 "chunks and record the outcome" if not has_index else
                 "exit code 2 while unresolved references remain; fix rules or correct "
                 "the variable names, then rerun"),
    }
    (workdir / "variable-verification.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    if args.summary:
        print(f"index: {result['index']['sheets']} sheets, "
              f"{result['index']['columns_total']} columns")
        print(f"references: {n_refs}  resolved: {result['resolved_count']}  "
              f"unresolved: {result['unresolved_count']}  "
              f"unverifiable: {result['unverifiable_count']}")
        for u in result["unresolved"][:20]:
            print(f"  UNRESOLVED [{u['rule_id']}] {u['ref']} — {u['problem']}"
                  + (f" (candidates: {', '.join(u['candidates'])})" if u["candidates"] else ""))
    else:
        json.dump({"references_total": n_refs,
                   "resolved": result["resolved_count"],
                   "unresolved": result["unresolved_count"],
                   "unverifiable": result["unverifiable_count"]},
                  sys.stdout, ensure_ascii=False, indent=2)
        print()

    return 2 if result["unresolved_count"] else 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except SystemExit:
        raise
    except Exception as exc:
        fail(f"{type(exc).__name__}: {exc}")
